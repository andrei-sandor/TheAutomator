import pandas
import matplotlib
import pandas as pd
import numpy
import win32com.client
from openpyxl import *
import random


#TODO add export file
#TODO add advanced mode
def excel_automator(name_of_file, average_columns,min_columns, max_columns, std_columns, refresh, advanced_mode, plot):
    df_excel = pd.read_excel(str(name_of_file))

    if average_columns:
        for column in df_excel:
            df_excel[column] = numpy.average(df_excel[column])

    if min_columns:
        for column in df_excel:
            df_excel["Min for ", column] = numpy.min(df_excel[column])

    if max_columns:
        for column in df_excel:
            df_excel["Max for", column] = numpy.max(df_excel[column])

    if std_columns:
        for column in df_excel:
            df_excel["std for", column] = numpy.std(df_excel)

    if refresh:
        excel_file = win32com.client.Dispatch(str(name_of_file))
        excel_file.Visible = 1
        file_open = excel_file.Workbooks.open(workbook)
        file_open.RefreshAll()
        file_open.Save()
        excel_file.Quit()

    if advanced_mode:
        excel = load_workbook(str(name_of_file)) #wb
        excel_active = excel.active #ws

        excel_sheet = excel["Sheet1"]
        excel_sheet.title = "New title"
        excel.save(str(name_of_file))

        excel_sheet = excel["Sheet name"]
        excel_sheet.sheet_properties.tabColor = "red"
        excel.save(str(name_of_file))

        excel = openpyxl.load_workbook(str(name_of_file))
        excel_active = excel.active
        number_cell = random.randint(0,10)
        value = excel_active["A" + str(number_cell)].value

        excel_active["A"+ str(number_cell)].value = "Hello"
        excel.save(str(name_of_file))

        excel_active.merge_cells("A10:B10")
        excel.save(str(name_of_file))

        excel_active.cell(row=10,column=10).value = "New value"
        excel_active.cell(row=10,column=11).value = "New value 2"

        excel_active.row_dimensions[1].height = 100
        excel_active.column_dimensions["A"].width = 100
        excel.save(str(name_of_file))

        excel_active.insert_rows(10)
        excel_active.insert_cols(10)
        excel.save(str(name_of_file))

        excel_active.cell(rows=10,column=10).value = "Hello"
        excel_active.cell(rows=10,column=10).font = Font(size = 18)

        excel.save(str(name_of_file))

        color_cell = PatternFill(patternType="solid", fgColor="Red")
        excel_active["A1"].fill = color_cell
        excel.save(str(name_of_file))

        row = 10
        column = 10

        array_insert = ["A", "B", "C"]

        for element in array_insert:
            excel_active.write(row,column,element)
            row+=2

        excel.save(str(name_of_file))

        excel_active["Z1"] = '=SUM("A1:A10)'
        excel_active["Z1"] = '=PRODUCT("A1:A10)'
        excel_active["Z1"] = '=COUNT("A1:A10)'
        excel_active["Z1"] = '=PRODUCT("A1:A10)'

    if plot:
        xpoints = numpy.array([1,10])
        ypoints = numpy.array(df_excel.loc[0:11,"A1"])

        matplotlib.plot(xpoints,ypoints, 'o')
        matplotlib.show()
